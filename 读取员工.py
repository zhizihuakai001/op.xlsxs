from openpyxl.reader.excel import load_workbook

wb2= load_workbook(filename='缝盘2组6月份.xlsx', data_only= True)
excel = wb2.active # 当前激活的工作表
for i in range(100,2):
    value= excel.cell(row = i).value
    print(value)