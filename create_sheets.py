from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

wb = Workbook() # 开始工作
ws = wb.active # 一个工作表至少有一个工作簿

ws1=wb.create_sheet('mysheet1',1) # 在末尾插入(默认)
ws2=wb.create_sheet('mysheet2', 0) # 插入第一个位置
ws3 = wb.create_sheet('mysheet0',-1) # 插入倒数第二个位置
ws.title='new title' # 工作薄在创建时会自动生成一个名字，以(Sheet, Sheet1, Sheet2, …)来进行命名。你也可以通过 Worksheet.title 属性来修改命名:
#ws3= wb['new title']
print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)
# c = ws['A4']
# ws['A4'] = 4
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

    tuple(ws.rows)
for row in ws.values:
    for value in row:
        print(value)
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    print(row)
wb.save('bubu.xlsx')

wb2= load_workbook(filename='缝盘2组6月份.xlsx', data_only= True)
excel = wb2.active # 当前激活的工作表
value= excel.cell(row = 4, column= 1).value
print(value)
#print(list(excel.values))
print(excel.max_row) #最大列数
print(excel.max_column) #最大行数